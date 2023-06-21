import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import random
from datetime import time

def generate_clock_table():
    employees = [f"员工{i}" for i in range(1, 51)]
    dates = pd.date_range(start="2023-05-01", end="2023-05-30")
    data = {date: [generate_clock_time() for _ in range(50)] for date in dates}
    df = pd.DataFrame(data, index=employees)
    df.to_excel("clock.xlsx", engine="openpyxl")

def generate_clock_time():
    num_times = np.random.choice([2, 3, 4, 1, 0], p=[0.8, 0.05, 0.05, 0.08, 0.02])
    times = []
    for _ in range(num_times):
        if len(times) == 0:
            first_time = random.uniform(9*60 + 30, 10*60 + 20)
            times.append(time(hour=int(first_time // 60), minute=int(first_time % 60)))
        elif len(times) == num_times - 1:
            last_time = random.uniform(18*60 + 50, 20*60)
            times.append(time(hour=int(last_time // 60), minute=int(last_time % 60)))
        else:
            times.append(time(hour=random.randint(10, 18), minute=random.randint(0, 59)))
    times.sort()
    return "\n".join([t.strftime("%H:%M") for t in times])

def process_clock_table():
    df = pd.read_excel("clock.xlsx", engine="openpyxl")
    late_counts = []
    early_leaves = []

    wb = load_workbook("clock.xlsx")
    ws = wb.active
    red_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_idx, row in df.iterrows():
        late_count = 0
        early_leave = 0
        for col_idx, cell in enumerate(row):
            times = cell.split("\n") if isinstance(cell, str) else []
            if len(times) > 1:
                first_time = times[0]
                last_time = times[-1]
                if first_time > "10:00":
                    late_count += 1
                    ws.cell(row=row_idx + 2, column=col_idx + 2).fill = red_fill
                if last_time < "19:30":
                    early_leave += 1
                    ws.cell(row=row_idx + 2, column=col_idx + 2).fill = red_fill

        late_counts.append(late_count)
        early_leaves.append(early_leave)

    df["迟到次数"] = late_counts
    df["早退次数"] = early_leaves
    late_people = df.apply(lambda x: sum([1 for v in x[:-2] if isinstance(v, str) and v.split("\n")[0] > "10:00"]), axis=0)
    early_leave_people = df.apply(lambda x: sum([1 for v in x[:-2] if isinstance(v, str) and v.split("\n")[-1] < "19:30"]), axis=0)
    df.loc["迟到人数"] = late_people
    df.loc["早退人数"] = early_leave_people

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    wb.save("clock_processed.xlsx")

if __name__ == "__main__":
    generate_clock_table()
    process_clock_table()